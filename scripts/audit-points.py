import openpyxl

wb = openpyxl.load_workbook('../Audit_Website 2025.xlsx', data_only=True)
ws = wb['Audit Point']

items = []
for r in range(2, ws.max_row + 1):
    name = ws.cell(r, 1).value
    status = ws.cell(r, 2).value
    tool = ws.cell(r, 3).value
    guide = ws.cell(r, 4).value
    if name:
        items.append((r, name, status, tool, guide))

print(f'Total audit items: {len(items)}')

true_count = sum(1 for _, _, s, _, _ in items if s == True)
false_count = sum(1 for _, _, s, _, _ in items if s == False)
other = sum(1 for _, _, s, _, _ in items if s not in (True, False, None))
none_status = sum(1 for _, _, s, _, _ in items if s is None)
cat = sum(1 for _, _, s, t, g in items if s is None and t is None and g is None)

print(f'Categories (sections): {cat}')
print(f'Passed (True): {true_count}')
print(f'Failed (False): {false_count}')
print(f'Other: {other}')
print(f'None (unchecked): {none_status}')
print()

for idx, (r, name, status, tool, guide) in enumerate(items, 1):
    is_cat = status is None and tool is None and guide is None
    st = '' if status is None else ('OK' if status == True else 'FAIL')
    marker = '--- ' if is_cat else '     '
    g = (guide or '')[:60].replace('\n', ' ')
    t = (tool or '').replace('\n', ' ')
    print(f'{r:4d} {marker}{st:4s} | {name}')
    if t:
        print(f'         Tool: {t}')
    if g:
        print(f'         Guide: {g}')
